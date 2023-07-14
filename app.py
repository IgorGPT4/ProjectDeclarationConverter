import ipywidgets as widgets
from IPython.display import display
import requests
import pdfplumber
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def pdf_excel(pdf_url, factors_list, excel_file):
    # Отправка GET-запроса для загрузки файла
    response = requests.get(pdf_url)

    # Проверка статуса ответа
    if response.status_code == 200:
        # Получение содержимого файла
        pdf_content = response.content

        # Открытие PDF-файла с помощью pdfplumber
        with pdfplumber.open(BytesIO(pdf_content)) as pdf:
            # Создание нового Excel-файла
            workbook = Workbook()
            sheet = workbook.active

            # Переменная для хранения номера текущей строки
            current_row = 1

            # Извлечение текстовых блоков из первой страницы
            first_page = pdf.pages[0]
            text_blocks = first_page.extract_text().split("\n")

            # Запись второй строки первой страницы в Excel
            if len(text_blocks) >= 2:
                extracted_text = text_blocks[1]
                # Форматирование текста
                formatted_text = ""
                for index, char in enumerate(extracted_text):
                    if char != " " and index % 2 == 0:
                        continue
                    formatted_text += char
                sheet.cell(row=current_row, column=2).value = formatted_text
                current_row += 1

            # Итерация по каждой странице
            for page in pdf.pages:
                # Извлечение текстовых блоков из текущей страницы
                text_blocks = page.extract_text().split("\n")

                # Проверка наличия текстового блока с текстом "Объект №"
                for text_block in text_blocks:
                    if "ООббъъеекктт №№" in text_block:
                        formatted_text = ""
                        for index, char in enumerate(text_block):
                            if char != " " and index % 2 == 0:
                                continue
                            formatted_text += char

                        sheet.cell(row=current_row, column=2).value = formatted_text
                        current_row += 1

                # Извлечение таблицы из страницы
                table = page.extract_table()

                # Проверка наличия данных в таблице
                if table:
                    rows = len(table)
                    cols = len(table[0])

                    # Запись таблицы в Excel-файл
                    for i in range(rows):
                        if table[i][1] and table[i][1] in factors_list and table[i][2] is not None:
                            # Найти индекс первого встречающегося символа ":"
                            colon_index = table[i][2].find(":")
                            if colon_index != -1:
                                # Извлечь текст после символа ":"
                                extracted_text = table[i][2][colon_index + 1:].strip()

                                # Форматирование текста
                                formatted_text = ""
                                for index, char in enumerate(extracted_text):
                                    if char != " " and index % 2 == 0:
                                        continue
                                    formatted_text += char

                                sheet.cell(row=current_row, column=4).value = formatted_text
                            else:
                                sheet.cell(row=current_row, column=4).value = ""

                            for j in range(cols):
                                sheet.cell(row=current_row, column=j + 1).value = table[i][j]
                            current_row += 1

                # Извлечение текстовых блоков из текущей страницы
                text_blocks = page.extract_text().split("\n")

                # Запись строк с текстом "Объект №" в Excel
                for text_block in text_blocks:
                    if "Объект №" in text_block:
                        current_row += 1
                        sheet.cell(row=current_row, column=2).value = text_block

            # Сохранение Excel-файла
            workbook.save(excel_file)
            return workbook
    else:
        print("Ошибка при загрузке PDF-файла")

def in_excel(workbook):
    # Получение активного листа
    sheet = workbook.active

    # Получение количества строк и столбцов в листе
    rows = sheet.max_row
    cols = sheet.max_column

    # Удаление первого столбца
    sheet.delete_cols(1)

    # Обработка ширины столбцов
    sheet.column_dimensions[get_column_letter(1)].width = 25.00
    sheet.column_dimensions[get_column_letter(2)].width = 130.00
    sheet.column_dimensions[get_column_letter(3)].width = 25.00

    # Обработка ячеек в третьем столбце
    for row in range(1, rows + 1):
        cell_value = sheet.cell(row=row, column=2).value
        if cell_value is not None:
            colon_index = cell_value.find(':')
            if colon_index != -1:
                modified_value = cell_value[:colon_index]
                sheet.cell(row=row, column=2).value = modified_value

    return workbook

def process_pdf(pdf_url, factors, excel_file, output_widget):
    with output_widget:
        print("Выполняется конвертация PDF...")
    workbook = pdf_excel(pdf_url, factors, excel_file)
    workbook2 = in_excel(workbook)
    workbook2.save(excel_file)
    with output_widget:
        print("PDF успешно обработан. Файл сохранен:", excel_file)

# Создание виджетов для ввода данных
pdf_url_widget = widgets.Text(description='URL PDF-файла:')
factors_widget = widgets.Text(description='Факторы (через запятую):', value='4.1.11, 9.3.1, 9.3.2, 12.3.2, 9.2.21, 18.1.1, 19.6.1.4, 19.6.1.6, 19.7.2.1.1.1, 19.7.3.1.1.1')
excel_file_widget = widgets.Text(description='Имя Excel-файла:', value='output.xlsx')
convert_button = widgets.Button(description='Конвертировать')
output_widget = widgets.Output()

# Функция-обработчик события нажатия на кнопку
def on_convert_button_clicked(b):
    pdf_url = pdf_url_widget.value
    factors = factors_widget.value.split(', ')
    excel_file = excel_file_widget.value

    if pdf_url and factors and excel_file:
        process_pdf(pdf_url, factors, excel_file, output_widget)
    else:
        with output_widget:
            print("Пожалуйста, заполните все поля.")

# Привязка функции-обработчика к событию нажатия на кнопку
convert_button.on_click(on_convert_button_clicked)

# Отображение виджетов
display(pdf_url_widget, factors_widget, excel_file_widget, convert_button, output_widget)